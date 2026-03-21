"""Tab 2: Drawing Analyser & QTO."""

import pathlib
import time
import streamlit as st
from utils import gemini_client
from utils.i18n import t, t_fmt
from utils.ui_components import section_header, metric_row

DEMO_DIR = pathlib.Path(__file__).resolve().parent.parent / "assets" / "demo_drawings"

def _analysis_focus() -> dict[str, str]:
    """Build ANALYSIS_FOCUS at call-time so t() picks up the active locale."""
    return {
        t("af_qto"): "QTO",
        t("af_discrepancy"): "ERRORS",
        t("af_compliance"): "COMPLIANCE",
        t("af_constructability"): "CONSTRUCTABILITY",
        t("af_cost"): "COST",
    }


def _progress_messages() -> list[str]:
    """Build progress messages at call-time so t() picks up the active locale."""
    return [
        t("t2_p1"),
        t("t2_p2"),
        t("t2_p3"),
        t("t2_p4"),
        t("t2_p5"),
        t("t2_p6"),
        t("t2_p7"),
    ]


# ---------------------------------------------------------------------------
# Prompt builder
# ---------------------------------------------------------------------------

def _build_analysis_prompt(focus_keys: list[str]) -> str:
    """Build the Gemini prompt based on selected analysis focus and region."""
    from utils.i18n import get_region
    is_cn = get_region() == "cn"

    qto_block = """
## QUANTITY TAKE-OFF

| Item | Description | Quantity | Unit | Measurement Method | Notes |
|------|-------------|----------|------|--------------------|-------|
[Extract EVERY measurable item. Include: concrete (m³), reinforcement (t), earthworks (m³), pipes (m), pits (nr), kerbs (m), pavements (m²), landscaping (m²), fencing (m), etc.]
[Read dimensions directly from the drawing. Show calculations where relevant.]

### QTO Summary
- Total line items extracted: X
- Key cost drivers identified: [list]
- Items requiring field verification: [list]
"""

    errors_block = """
## DISCREPANCIES & ERRORS

| # | Type | Location | Description | Severity | Recommendation |
|---|------|----------|-------------|----------|----------------|
[Check for: missing dimensions, scale inconsistencies, clashing elements, incomplete annotations, non-standard symbols, drainage fall issues, clearance violations, missing references, title block errors, north point issues]
Use severity indicators: 🔴 Critical, 🟡 Warning, 🟢 Minor

### Error Summary
- 🔴 Critical (must fix before issue): X
- 🟡 Warning (should review): Y
- 🟢 Minor (note for record): Z
"""

    if is_cn:
        compliance_block = """
## COMPLIANCE OBSERVATIONS

| Standard | Clause | Status | Observation |
|----------|--------|--------|-------------|
[Check against: GB 50011 (seismic), GB 50015 (plumbing/drainage), JGJ 100 (parking), GB 50763 (accessibility), GB 50016 (fire protection), local planning regulations as applicable]
Use status: ✅ Pass, ⚠️ Flag, ❌ Fail
"""
        cost_block = """
## COST ESTIMATION INDICATORS

| Item | Estimated Quantity | Unit Rate Range (RMB) | Estimated Cost Range | Notes |
|------|-------------------|----------------------|---------------------|-------|
[Based on extracted quantities, provide preliminary cost ranges using typical Chinese civil construction rates per GB 50500.]

### Budget Flags
[Identify items that may significantly impact budget or have high cost uncertainty.]
"""
    else:
        compliance_block = """
## COMPLIANCE OBSERVATIONS

| Standard | Clause | Status | Observation |
|----------|--------|--------|-------------|
[Check against: BCA, AS 3500 (plumbing/drainage), AS 2890 (parking), AS 1428 (access), local council DCPs as applicable]
Use status: ✅ Pass, ⚠️ Flag, ❌ Fail
"""
        cost_block = """
## COST ESTIMATION INDICATORS

| Item | Estimated Quantity | Unit Rate Range (AUD) | Estimated Cost Range | Notes |
|------|-------------------|----------------------|---------------------|-------|
[Based on extracted quantities, provide preliminary cost ranges using typical Australian civil construction rates.]

### Budget Flags
[Identify items that may significantly impact budget or have high cost uncertainty.]
"""

    constructability_block = """
## CONSTRUCTABILITY NOTES
[Practical observations about construction sequence, access for plant, staging, potential clashes with existing services, temporary works requirements, and any buildability concerns.]
"""

    # Assemble the sections based on selected focus
    sections = []
    if "QTO" in focus_keys:
        sections.append(qto_block)
    if "ERRORS" in focus_keys:
        sections.append(errors_block)
    if "COMPLIANCE" in focus_keys:
        sections.append(compliance_block)
    if "CONSTRUCTABILITY" in focus_keys:
        sections.append(constructability_block)
    if "COST" in focus_keys:
        sections.append(cost_block)

    focus_desc = ", ".join(focus_keys)

    if is_cn:
        role = "a senior civil drafter and quantity surveyor at a leading Chinese consulting engineering firm"
    else:
        role = "a senior civil drafter and quantity surveyor at a leading Australian consulting engineering firm"

    return f"""You are {role}. Analyse this construction drawing in detail.

First identify the drawing type (e.g. Site Plan, Floor Plan, Structural Plan, Elevation, Section, Services, Landscape, etc.) from the drawing content.
Analysis Required: {focus_desc}

PROVIDE YOUR ANALYSIS IN THIS EXACT STRUCTURE:

## DRAWING OVERVIEW
- Drawing title/number (if visible)
- Scale (if shown)
- Overall description of what is depicted
- Key standards/references noted

{"".join(sections)}

Be thorough and specific. Reference exact locations on the drawing (e.g., "grid line B-3", "north-east corner", "chainage 45m"). This analysis will be used by the design team to improve documentation quality."""


# ---------------------------------------------------------------------------
# Demo helpers
# ---------------------------------------------------------------------------

def _get_demo_drawings() -> list[pathlib.Path]:
    if DEMO_DIR.exists():
        return sorted(
            p for p in DEMO_DIR.iterdir()
            if p.suffix.lower() in {".jpg", ".jpeg", ".png", ".pdf"}
        )
    return []


# ---------------------------------------------------------------------------
# Render
# ---------------------------------------------------------------------------

def render():

    # --- Layout -------------------------------------------------
    col_left, col_right = st.columns([2, 3])

    image_bytes: bytes | None = None
    mime_type = "image/jpeg"

    with col_left:
        st.markdown(f"##### {t('t2_drawing')}")
        uploaded = st.file_uploader(
            t("t2_upload"),
            type=["jpg", "jpeg", "png", "pdf"],
            key="tab2_upload",
        )

        if uploaded is not None:
            image_bytes = uploaded.getvalue()
            mime_type = uploaded.type or "image/jpeg"
            if uploaded.type and "pdf" in uploaded.type:
                st.info(f"📄 PDF uploaded: **{uploaded.name}** ({len(image_bytes) / 1024:.0f} KB)")
            else:
                st.image(image_bytes, caption="Uploaded Drawing", use_container_width=True)

        demo_drawings = _get_demo_drawings()
        if False and demo_drawings:
            st.markdown(
                f"<span style='color:#8B949E;font-size:0.8rem;'>{t('t2_demo_label')}</span>",
                unsafe_allow_html=True,
            )
            demo_cols = st.columns(min(len(demo_drawings), 3))
            for i, drawing in enumerate(demo_drawings):
                with demo_cols[i % 3]:
                    if st.button(drawing.stem, key=f"demo_drawing_{i}", use_container_width=True):
                        st.session_state["tab2_demo_drawing"] = str(drawing)

            if "tab2_demo_drawing" in st.session_state:
                demo_path = pathlib.Path(st.session_state["tab2_demo_drawing"])
                if demo_path.exists():
                    image_bytes = demo_path.read_bytes()
                    if demo_path.suffix.lower() == ".pdf":
                        mime_type = "application/pdf"
                        st.info(f"📄 Demo PDF: **{demo_path.stem}**")
                    else:
                        mime_type = "image/png" if demo_path.suffix.lower() == ".png" else "image/jpeg"
                        st.image(image_bytes, caption=f"Demo — {demo_path.stem}", use_container_width=True)

    with col_right:
        st.markdown(f"##### {t('t2_analysis_params')}")

        st.markdown(
            f"<span style='color:#8B949E;font-size:0.8rem;letter-spacing:0.5px;'>"
            f"{t('t2_focus_label')}</span>",
            unsafe_allow_html=True,
        )
        focus_map = _analysis_focus()
        selected_focus = st.multiselect(
            t("t2_focus"),
            list(focus_map.keys()),
            default=[t("af_qto"), t("af_discrepancy")],
            key="tab2_focus",
            label_visibility="collapsed",
        )

        analyse = st.button(
            t("t2_analyse_btn"),
            type="primary",
            use_container_width=True,
            disabled=image_bytes is None or len(selected_focus) == 0,
        )

        if image_bytes is None:
            st.info(t("t2_upload_hint"))
        elif len(selected_focus) == 0:
            st.warning(t("t2_select_focus"))

    # --- Analysis execution -------------------------------------
    if analyse and image_bytes is not None and selected_focus:
        from utils.auth import get_user
        from utils.usage import check_quota, render_quota_exceeded, record_usage
        _user = get_user()
        if _user:
            _allowed, _, _ = check_quota(_user["id"])
            if not _allowed:
                render_quota_exceeded()
                st.stop()
        st.markdown("---")

        focus_keys = [focus_map[f] for f in selected_focus]
        prompt = _build_analysis_prompt(focus_keys)

        # --- Custom progress bar ---
        progress_container = st.empty()
        status_text = st.empty()

        def _render_progress(pct: int, msg: str, done: bool = False):
            bar_color = "#3FB950" if done else "#FF6B35"
            icon_html = "" if done else f"""
                <div style="
                    position:absolute;
                    left:calc({pct}% - 18px); top:-14px;
                    font-size:20px;
                    transition: left 0.8s ease-in-out;
                    filter: drop-shadow(0 0 4px rgba(255,107,53,0.6));
                ">📋</div>"""
            check_html = """
                <div style="position:absolute;right:4px;top:-16px;font-size:18px;">✅</div>""" if done else ""

            progress_container.markdown(f"""
                <div style="
                    position:relative; background:#21262D; border-radius:8px;
                    height:10px; margin:22px 0 8px 0; overflow:visible; border:1px solid #30363D;
                ">
                    <div style="
                        width:{pct}%; height:100%; border-radius:8px;
                        background:linear-gradient(90deg, {bar_color}88, {bar_color});
                        transition:width 0.8s ease-in-out;
                        box-shadow:0 0 8px {bar_color}66;
                    "></div>
                    {icon_html}{check_html}
                </div>
            """, unsafe_allow_html=True)

            if done:
                status_text.markdown(
                    f"<div style='display:flex;align-items:center;gap:8px;'>"
                    f"<span style='color:#3FB950;font-weight:700;'>{t('t2_complete')}</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            else:
                status_text.markdown(
                    f"<div style='display:flex;align-items:center;gap:8px;'>"
                    f"<span style='color:#FF6B35;font-weight:600;'>🔬 Analysing</span>"
                    f"<span style='color:#8B949E;'>|</span>"
                    f"<span style='color:#E6EDF3;'>{msg}</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

        # Animate progress before API call
        progress_msgs = _progress_messages()
        _render_progress(1, progress_msgs[0])
        time.sleep(0.4)

        for i, msg in enumerate(progress_msgs):
            pct = int(((i + 1) / len(progress_msgs)) * 60) + 1
            _render_progress(pct, msg)
            time.sleep(0.5)

        _render_progress(65, "Waiting for ArchiMind Pro response...")

        # API call
        result_text = gemini_client.analyse_image(
            prompt=prompt,
            image_bytes=image_bytes,
            mime_type=mime_type,
        )

        if result_text:
            _render_progress(100, "", done=True)
            st.session_state["tab2_results"] = result_text
            st.session_state["tab2_focus_keys"] = focus_keys
            # Record usage only after successful analysis
            if _user:
                record_usage(_user["id"], "drawing_analyser", gemini_client.get_model_id())
                # Save to history
                try:
                    from utils.history import save_history
                    save_history(
                        _user["id"], "drawing_analyser",
                        f"Drawing Analysis — {', '.join(focus_keys)}",
                        result_text[:500],
                    )
                except Exception:
                    pass
                st.rerun()
        else:
            progress_container.empty()
            status_text.empty()
            st.error("Analysis failed. Please adjust parameters and retry.")

    # --- Display results ----------------------------------------
    if "tab2_results" in st.session_state and st.session_state["tab2_results"]:
        result_text = st.session_state["tab2_results"]
        focus_keys = st.session_state.get("tab2_focus_keys", [])

        st.markdown("---")
        section_header("📊", t("t2_results_title"))

        # Build sub-tabs based on what was analysed
        tab_labels = [t("t2_tab_overview")]
        if "QTO" in focus_keys:
            tab_labels.append(t("t2_tab_qto"))
        if "ERRORS" in focus_keys:
            tab_labels.append(t("t2_tab_discrepancies"))
        if "COMPLIANCE" in focus_keys:
            tab_labels.append(t("t2_tab_compliance"))
        if "CONSTRUCTABILITY" in focus_keys:
            tab_labels.append(t("t2_tab_constructability"))
        if "COST" in focus_keys:
            tab_labels.append(t("t2_tab_cost"))

        # Parse the result into sections
        sections = _parse_sections(result_text)

        result_tabs = st.tabs(tab_labels)

        tab_idx = 0

        # Overview tab (always shown)
        with result_tabs[tab_idx]:
            overview = sections.get("DRAWING OVERVIEW", "")
            if overview:
                st.markdown(overview)
            else:
                st.markdown(result_text[:500] + "..." if len(result_text) > 500 else result_text)
        tab_idx += 1

        # QTO tab
        if "QTO" in focus_keys:
            with result_tabs[tab_idx]:
                qto = sections.get("QUANTITY TAKE-OFF", "")
                if qto:
                    st.markdown(qto)
                else:
                    st.info(t("t2_no_qto"))
            tab_idx += 1

        # Errors tab
        if "ERRORS" in focus_keys:
            with result_tabs[tab_idx]:
                errors = sections.get("DISCREPANCIES & ERRORS", "") or sections.get("DISCREPANCIES", "")
                if errors:
                    st.markdown(errors)
                else:
                    st.info(t("t2_no_discrepancies"))
            tab_idx += 1

        # Compliance tab
        if "COMPLIANCE" in focus_keys:
            with result_tabs[tab_idx]:
                compliance = sections.get("COMPLIANCE OBSERVATIONS", "") or sections.get("COMPLIANCE", "")
                if compliance:
                    st.markdown(compliance)
                else:
                    st.info(t("t2_no_compliance"))
            tab_idx += 1

        # Constructability tab
        if "CONSTRUCTABILITY" in focus_keys:
            with result_tabs[tab_idx]:
                construct = sections.get("CONSTRUCTABILITY NOTES", "") or sections.get("CONSTRUCTABILITY", "")
                if construct:
                    st.markdown(construct)
                else:
                    st.info(t("t2_no_constructability"))
            tab_idx += 1

        # Cost tab
        if "COST" in focus_keys:
            with result_tabs[tab_idx]:
                cost = sections.get("COST ESTIMATION INDICATORS", "") or sections.get("COST", "")
                if cost:
                    st.markdown(cost)
                else:
                    st.info(t("t2_no_cost"))
            tab_idx += 1

        # --- Downloads ---
        st.markdown("---")
        dl_cols = st.columns(3)
        with dl_cols[0]:
            st.download_button(
                label=t("t2_download_report"),
                data=result_text,
                file_name="drawing_analysis_report.md",
                mime="text/markdown",
                key="dl_tab2_report",
                use_container_width=True,
            )
        with dl_cols[1]:
            # Extract just the QTO table if present
            qto_csv = _extract_qto_csv(result_text)
            if qto_csv:
                st.download_button(
                    label=t("t2_download_qto"),
                    data=qto_csv,
                    file_name="quantity_takeoff.csv",
                    mime="text/csv",
                    key="dl_tab2_csv",
                    use_container_width=True,
                )
        with dl_cols[2]:
            # Excel export of QTO
            excel_data = _generate_qto_excel(result_text)
            if excel_data:
                st.download_button(
                    label="📊 Excel (.xlsx)",
                    data=excel_data,
                    file_name="quantity_takeoff.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_tab2_xlsx",
                    use_container_width=True,
                )

        # --- Feedback ------------------------------------------------
        from utils.feedback import render_feedback
        render_feedback("drawing_analyser", "tab2_results")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_sections(text: str) -> dict[str, str]:
    """Split the Gemini response into named sections keyed by ## headings."""
    sections: dict[str, str] = {}
    current_key = ""
    current_lines: list[str] = []

    for line in text.split("\n"):
        if line.startswith("## "):
            if current_key:
                sections[current_key] = "\n".join(current_lines).strip()
            current_key = line[3:].strip()
            current_lines = []
        else:
            current_lines.append(line)

    if current_key:
        sections[current_key] = "\n".join(current_lines).strip()

    return sections


def _extract_qto_csv(text: str) -> str | None:
    """Extract the first markdown table from the QTO section and convert to CSV."""
    sections = _parse_sections(text)
    qto = sections.get("QUANTITY TAKE-OFF", "")
    if not qto:
        return None

    csv_lines = []
    for line in qto.split("\n"):
        stripped = line.strip()
        if stripped.startswith("|") and stripped.endswith("|"):
            # Skip separator rows (|---|---|...)
            cells = [c.strip() for c in stripped.split("|")[1:-1]]
            if cells and not all(set(c) <= {"-", ":", " "} for c in cells):
                csv_lines.append(",".join(f'"{c}"' for c in cells))

    if len(csv_lines) <= 1:
        return None

    return "\n".join(csv_lines)


def _generate_qto_excel(text: str) -> bytes | None:
    """Generate an Excel file from the QTO section."""
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        sections = _parse_sections(text)
        qto = sections.get("QUANTITY TAKE-OFF", "")
        if not qto:
            return None

        # Parse table rows
        rows = []
        for line in qto.split("\n"):
            stripped = line.strip()
            if stripped.startswith("|") and stripped.endswith("|"):
                cells = [c.strip() for c in stripped.split("|")[1:-1]]
                if cells and not all(set(c) <= {"-", ":", " "} for c in cells):
                    rows.append(cells)

        if len(rows) <= 1:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Quantity Take-Off"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="0A7CFF", end_color="0A7CFF", fill_type="solid")
        border = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )

        # Write header
        for col, val in enumerate(rows[0], 1):
            cell = ws.cell(row=1, column=col, value=val)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Write data
        for row_idx, row_data in enumerate(rows[1:], 2):
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = border

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # Title row above table
        ws.insert_rows(1)
        title_cell = ws.cell(row=1, column=1, value="ArchiMind Pro — Quantity Take-Off Report")
        title_cell.font = Font(bold=True, size=14, color="0A7CFF")

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except Exception:
        return None
